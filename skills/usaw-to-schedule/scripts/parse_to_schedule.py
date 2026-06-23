#!/usr/bin/env python3
"""Parse a USAW TO Sign-up workbook tab: list day separators, sessions,
platforms, role assignments — or look up one person's assignments across tabs.

No pip in this env -> run via: uv run --with openpyxl python parse_to_schedule.py ...

Usage:
  parse_to_schedule.py <xlsx> --tab "2026 NCW"
  parse_to_schedule.py <xlsx> --person "Kelly Wiese"
"""
import sys, re, argparse, datetime

ROLE_COLS_DEFAULT = {8: "Weigh-in", 9: "Speaker", 10: "Timekeeper",
                     11: "Referee", 12: "TC", 13: "Marshal", 14: "Jury"}
MONTHS = {m.upper(): i for i, m in enumerate(
    ["", "January", "February", "March", "April", "May", "June", "July",
     "August", "September", "October", "November", "December"])}


def norm(v):
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def find_header_row(ws):
    for r in range(1, 20):
        if norm(ws.cell(r, 1).value) == "Session":
            return r
    return 7


def parse_day(s):
    m = re.search(r"(\d{1,2}),\s*(\d{4})", s)
    mo = re.search(r"(" + "|".join(k for k in MONTHS if k) + r")", s, re.I)
    if m and mo:
        return datetime.date(int(m.group(2)), MONTHS[mo.group(1).upper()], int(m.group(1)))
    return None


def iter_assignments(ws, roles=ROLE_COLS_DEFAULT):
    anchors = {}
    cur_day = None
    cur_sess = None
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and re.search(r"DAY,", a):
            cur_day = parse_day(a)
            continue
        if isinstance(a, (int, float)):
            cur_sess = int(a)
        plat = ws.cell(r, 2).value
        if plat in ("RED", "WHITE", "BLUE"):
            anchors[r] = dict(sess=cur_sess, plat=plat, day=cur_day,
                              gndr=ws.cell(r, 5).value, cat=norm(ws.cell(r, 6).value),
                              win=ws.cell(r, 4).value, start=ws.cell(r + 1, 4).value)

    def block_for(row):
        for rr in range(row, row - 4, -1):
            if rr in anchors:
                return anchors[rr]
        return None

    for r in range(min(anchors) if anchors else 8, ws.max_row + 1):
        for col, role in roles.items():
            v = norm(ws.cell(r, col).value)
            if not v:
                continue
            b = block_for(r)
            if not b:
                continue
            yield dict(row=r, role=role, name=v, **b)


def main():
    import openpyxl
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--tab", default="2026 NCW")
    ap.add_argument("--person")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)

    if args.person:
        tabs = [t for t in wb.sheetnames if t != "List of TOs"]
        print(f"Assignments for {args.person}:")
        total = 0
        for tab in tabs:
            ws = wb[tab]
            hits = [a for a in iter_assignments(ws)
                    if args.person.lower() in a["name"].lower()]
            seen = set()
            for a in sorted(hits, key=lambda x: (str(x["day"]), x["sess"] or 0)):
                key = (a["day"], a["sess"], a["plat"], a["role"])
                if key in seen:
                    continue
                seen.add(key)
                d = a["day"].strftime("%a %b %d") if a["day"] else "??"
                # Report times by role:
                #   Weigh-in  → W.In time (2 hrs before start)
                #   Marshal   → Start - 30 min (fixed USAW policy)
                #   All others → Start time
                if a["role"] == "Weigh-in":
                    st = a["win"].strftime("%H:%M") if a["win"] else "--"
                    time_label = "w.in "
                elif a["role"] == "Marshal":
                    if a["start"]:
                        import datetime as _dt
                        _s = a["start"]
                        _dt_obj = _dt.datetime(2000, 1, 1, _s.hour, _s.minute) - _dt.timedelta(minutes=30)
                        st = _dt_obj.strftime("%H:%M")
                    else:
                        st = "--"
                    time_label = "rpt  "
                else:
                    st = a["start"].strftime("%H:%M") if a["start"] else "--"
                    time_label = "start"
                print(f"  [{tab}] {d} S{a['sess']} {a['plat']:<5} {a['role']:<10} "
                      f"{a['cat'][:22]:<22} {time_label} {st} MT")
                total += 1
        print(f"  total: {total}")
        return

    ws = wb[args.tab]
    hr = find_header_row(ws)
    print(f"Tab: {args.tab}  header row={hr}  rows={ws.max_row}")
    days = [(r, norm(ws.cell(r, 1).value)) for r in range(1, ws.max_row + 1)
            if isinstance(ws.cell(r, 1).value, str) and re.search(r"DAY,", ws.cell(r, 1).value)]
    print(f"\nDay separators ({len(days)}):")
    for r, t in days:
        print(f"  row {r}: {t}")
    sessions = sorted({(a["sess"], a["day"]) for a in iter_assignments(ws) if a["sess"]})
    print(f"\nSessions: {len(sessions)}")


if __name__ == "__main__":
    main()
